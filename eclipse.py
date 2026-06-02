"""
eclipse.py — Planificador eclipse solar total 12-ago-2026 desde Alboraya.

Uso:
    python eclipse.py routes              # Calcula distancias/tiempos OSRM
    python eclipse.py centerline          # Añade distancia a línea central
    python eclipse.py all                 # Ambos en orden
    python eclipse.py routes --xlsx       # + genera Excel
    python eclipse.py all --xlsx          # Todo + Excel

Salida siempre en output/eclipse_sitios_rutas.csv.
Con --xlsx genera también output/eclipse_sitios_rutas.xlsx.
"""

import argparse
import csv
import json
import math
import re
import sys
import time
import urllib.request
import urllib.error

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

ORIGIN_LAT = 39.5516
ORIGIN_LNG = -0.3373
OSRM_URL = "http://router.project-osrm.org/route/v1/driving/{lng1},{lat1};{lng2},{lat2}?overview=false"
OUTPUT_CSV  = "output/eclipse_sitios_rutas.csv"
OUTPUT_XLSX = "output/eclipse_sitios_rutas.xlsx"
SITES_FILE  = "sites.json"

# Línea central NASA (lat, lng decimales)
# Fuente: https://eclipse.gsfc.nasa.gov/SEpath/SEpath2001/SE2026Aug12Tpath.html
CENTERLINE = [
    (82.275, 112.487), (85.295, 104.215), (87.278, 81.525),
    (87.823, 33.000),  (86.835, -1.638),  (85.403, -15.182),
    (83.932, -21.187), (82.495, -24.272), (81.110, -25.992),
    (79.773, -26.982), (78.483, -27.540), (77.233, -27.825),
    (76.017, -27.928), (74.837, -27.905), (73.683, -27.788),
    (72.557, -27.603), (71.450, -27.362), (70.365, -27.078),
    (69.298, -26.760), (68.247, -26.410), (67.210, -26.032),
    (66.185, -25.630), (65.172, -25.205), (64.168, -24.757),
    (63.172, -24.287), (62.183, -23.793), (61.200, -23.277),
    (60.222, -22.737), (59.245, -22.170), (58.272, -21.573),
    (57.297, -20.947), (56.322, -20.287), (55.343, -19.588),
    (54.362, -18.847), (53.372, -18.057), (52.372, -17.212),
    (51.360, -16.303), (50.333, -15.317), (49.285, -14.238),
    (48.212, -13.048), (47.102, -11.715), (45.943, -10.190),
    (44.713, -8.398),  (43.372, -6.188),  (41.817, -3.185),
    (39.408, 2.950),
]

R = 6371.0

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def load_sites():
    with open(SITES_FILE, encoding="utf-8") as f:
        return json.load(f)

def normalize_dur(dur):
    """'1m 33s' | '51s' | '1m' -> '1m 33s' | '0m 51s' | '1m 00s'"""
    dur = dur.strip()
    m = re.match(r'(?:(\d+)m\s*)?(?:(\d+)s)?$', dur)
    if not m:
        return dur
    mins = int(m.group(1)) if m.group(1) else 0
    secs = int(m.group(2)) if m.group(2) else 0
    return f"{mins}m {secs:02d}s"

# ---------------------------------------------------------------------------
# OSRM
# ---------------------------------------------------------------------------

def get_route(dest_lat, dest_lng):
    url = OSRM_URL.format(
        lng1=ORIGIN_LNG, lat1=ORIGIN_LAT,
        lng2=dest_lng,   lat2=dest_lat
    )
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "eclipse-planner/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        route = data["routes"][0]
        dist_km = round(route["distance"] / 1000, 1)
        dur_min = round(route["duration"] / 60)
        h, m = divmod(dur_min, 60)
        return dist_km, f"{h:02d}h {m:02d}min"
    except Exception as e:
        return None, f"ERROR: {e}"

def cmd_routes():
    sites = load_sites()
    total = len(sites)
    rows = []
    for i, s in enumerate(sites):
        print(f"[{i+1}/{total}] {s['espacio']}...", end=" ", flush=True)
        dist_km, dur_str = get_route(s["lat"], s["lng"])
        print(f"{dist_km} km, {dur_str}")
        rows.append({
            "Espacio":            s["espacio"],
            "Municipio":          s["municipio"],
            "Tipo":               s["tipo"],
            "Capacidad":          s["capacidad"],
            "Parking":            s["parking"],
            "Inicio totalidad":   s["inicio_totalidad"],
            "Duración totalidad": normalize_dur(s["duracion_totalidad"]),
            "Lat":                s["lat"],
            "Lng":                s["lng"],
            "Dist_carretera_km":  dist_km if dist_km else "",
            "Tiempo_Maps":        dur_str,
        })
        time.sleep(0.3)

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"\nListo. Guardado en {OUTPUT_CSV}")

# ---------------------------------------------------------------------------
# Centerline
# ---------------------------------------------------------------------------

def haversine(lat1, lng1, lat2, lng2):
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng/2)**2
    return 2 * R * math.asin(math.sqrt(a))

def cross_track_distance(lat, lng, lat1, lng1, lat2, lng2):
    d13 = haversine(lat1, lng1, lat, lng) / R
    d12 = haversine(lat1, lng1, lat2, lng2) / R
    if d12 == 0:
        return haversine(lat1, lng1, lat, lng)

    def bearing(la1, lo1, la2, lo2):
        la1, lo1, la2, lo2 = map(math.radians, [la1, lo1, la2, lo2])
        dlo = lo2 - lo1
        x = math.sin(dlo) * math.cos(la2)
        y = math.cos(la1) * math.sin(la2) - math.sin(la1) * math.cos(la2) * math.cos(dlo)
        return math.atan2(x, y)

    t13 = bearing(lat1, lng1, lat, lng)
    t12 = bearing(lat1, lng1, lat2, lng2)
    dxt = math.asin(math.sin(d13) * math.sin(t13 - t12))
    dat = math.acos(max(-1, min(1, math.cos(d13) / math.cos(dxt))))

    if dat > d12 or math.cos(t13 - t12) < 0:
        return min(haversine(lat1, lng1, lat, lng), haversine(lat2, lng2, lat, lng))
    return abs(dxt) * R

def min_dist_to_centerline(lat, lng):
    return round(min(
        cross_track_distance(lat, lng, CENTERLINE[i][0], CENTERLINE[i][1],
                             CENTERLINE[i+1][0], CENTERLINE[i+1][1])
        for i in range(len(CENTERLINE) - 1)
    ), 1)

def cmd_centerline():
    with open(OUTPUT_CSV, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    for row in rows:
        try:
            lat, lng = float(row["Lat"]), float(row["Lng"])
            row["Dist_centerline_km"] = min_dist_to_centerline(lat, lng)
        except (KeyError, ValueError):
            row["Dist_centerline_km"] = ""

    fieldnames = list(rows[0].keys())
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)

    print(f"Listo. Columna Dist_centerline_km añadida a {OUTPUT_CSV}")

    sorted_rows = sorted(
        [r for r in rows if r["Dist_centerline_km"] != ""],
        key=lambda r: float(r["Dist_centerline_km"])
    )
    print("\nTop 10 más cercanos a la línea central:")
    print(f"{'Sitio':<45} {'Municipio':<30} {'Dist CL':>8}  {'Duración'}")
    print("-" * 95)
    for r in sorted_rows[:10]:
        print(f"{r['Espacio']:<45} {r['Municipio']:<30} {r['Dist_centerline_km']:>6} km  {r['Duración totalidad']}")

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

def cmd_xlsx():
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        print(f"ERROR: falta dependencia — {e}")
        print("Instala con: pip install openpyxl pandas")
        sys.exit(1)

    df = pd.read_csv(OUTPUT_CSV, encoding="utf-8")
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sitios")
        ws = writer.sheets["Sitios"]

        header_fill = PatternFill("solid", start_color="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Arial", size=10)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font

        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(df[col].fillna("").astype(str).map(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

    print(f"Guardado: {OUTPUT_XLSX}  ({len(df)} filas, {len(df.columns)} columnas)")

# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Planificador eclipse solar 12-ago-2026 desde Alboraya"
    )
    parser.add_argument(
        "command",
        choices=["routes", "centerline", "all"],
        help="routes: rutas OSRM | centerline: distancia a línea central | all: ambos"
    )
    parser.add_argument("--xlsx", action="store_true", help="Genera también el Excel")
    args = parser.parse_args()

    if args.command == "routes":
        cmd_routes()
    elif args.command == "centerline":
        cmd_centerline()
    elif args.command == "all":
        cmd_routes()
        cmd_centerline()

    if args.xlsx:
        cmd_xlsx()

if __name__ == "__main__":
    main()
