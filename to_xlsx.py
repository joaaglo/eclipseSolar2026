"""
to_xlsx.py — Convierte eclipse_sitios_rutas.csv a eclipse_sitios_rutas.xlsx
con cabeceras en negrita, columnas ajustadas y filtros activados.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

df = pd.read_csv("output/eclipse_sitios_rutas.csv", encoding="utf-8")

out = "output/eclipse_sitios_rutas.xlsx"
with pd.ExcelWriter(out, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Sitios")
    ws = writer.sheets["Sitios"]

    # Cabecera: negrita + fondo azul oscuro + texto blanco
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Fuente normal para datos
    data_font = Font(name="Arial", size=10)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font

    # Ancho de columnas según contenido
    for col_idx, col in enumerate(df.columns, 1):
        max_len = max(df[col].fillna("").astype(str).map(len).max(), len(col)) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)

    # Filtros en cabecera
    ws.auto_filter.ref = ws.dimensions

    # Fila fija (freeze pane)
    ws.freeze_panes = "A2"

print(f"Guardado: {out}  ({len(df)} filas, {len(df.columns)} columnas)")
